import openpyxl
import os

from ProcesoText import preprocess_text




# Obtiene la ruta absoluta del directorio actual
current_directory = os.path.dirname(__file__)

# Construye la ruta al archivo Excel en la carpeta 'dataset'
excel_file_path = os.path.join(current_directory, '..', 'dataset', 'games.xlsx')

# Carga el archivo Excel
wb = openpyxl.load_workbook(excel_file_path)

# Selecciona la hoja de trabajo
sheet = wb.active

# Obtiene el índice de la columna "About the game" si está presente
about_game_column_index = None
supported_languages_column_index = None

for cell in sheet[1]:  # Itera sobre la primera fila
    if cell.value == "About the game":
        about_game_column_index = cell.column
    elif cell.value == "Supported languages":
        supported_languages_column_index = cell.column

if about_game_column_index is not None and supported_languages_column_index is not None:
    # Itera sobre las filas y agrega las descripciones procesadas a la columna "Descripción procesada"
    sheet.cell(row=1, column=supported_languages_column_index + 1, value="Descripción procesada")

    for i, row in enumerate(sheet.iter_rows(min_row=2, min_col=about_game_column_index, values_only=True), start=2):
        description = row[0]  # Obtén la descripción del juego
        processed_description = preprocess_text(description)  # Procesa la descripción
        sheet.cell(row=i, column=supported_languages_column_index + 1, value=processed_description)

    # Guarda el archivo Excel con las descripciones procesadas
    preprocessed_excel_path = os.path.join(current_directory, '..', 'dataset', 'games_processed.xlsx')
    wb.save(preprocessed_excel_path)
    print("Se ha creado el archivo 'games_processed.xlsx' con las descripciones procesadas.")
else:
    print("No se encontró la columna 'About the game' o 'Supported languages' en el archivo Excel.")
    
